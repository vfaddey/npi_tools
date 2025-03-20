import openpyxl
import numpy as np
from pathlib import Path
from .data import WellProperty, SeamProperty, AuxiliaryProperty


def load_excel_data(file_path: str):
    """ Загружает данные из Excel и возвращает параметры """
    try:
        if not Path(file_path).is_file():
            raise FileNotFoundError(f"Файл {file_path} не найден.")

        wb = openpyxl.load_workbook(file_path, keep_vba=True, data_only=True, keep_links=True)
        ws = wb['Исходные данные']

        # Проверяем, что важные данные не пустые
        required_cells = ['C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9',
                          'C12', 'C13', 'C14', 'C15', 'C16', 'C19', 'C20', 'C21', 'C22', 'C23']
        for cell in required_cells:
            if ws[cell].value is None:
                raise ValueError(f"Отсутствует значение в ячейке {cell}")

        seams = SeamProperty(
            ws['C2'].value,  # Проницаемость пласта, м^2
            ws['C3'].value,  # Объемный коэффициент
            ws['C4'].value,  # Вязкость жидкости, Па*с
            ws['C5'].value,  # Толщина пласта, м
            ws['C6'].value,  # Радиус зоны дренирования, м
            ws['C7'].value,  # Размер зоны дренирования по оси X, м
            ws['C8'].value,  # Размер зоны дренирования по оси Y, м
            ws['C9'].value  # Начальное пластовое давление, Па
        )

        well = WellProperty(
            ws['C12'].value,  # Дебит скважины, м³/с
            ws['C13'].value,  # Радиус скважины, м
            ws['C14'].value,  # Половина длины трещины, м
            ws['C15'].value,  # Проницаемость трещины, м²
            ws['C16'].value,  # Ширина трещины, м
            None  # Доп. параметр, если нужен
        )

        aux = AuxiliaryProperty(
            ws['C19'].value,  # Начальная координата
            ws['C20'].value,  # Точность
            ws['C21'].value,  # Шаг вычисления притока
            ws['C22'].value,  # Проницаемость загрязненной зоны
            ws['C23'].value,  # Точка начала загрязненной зоны
            np.flip([cell.value for cell in ws['E'][1:21]]).tolist()  # Длина загрязненной зоны
        )

        return seams, well, aux

    except Exception as e:
        print(f"Ошибка при загрузке данных: {e}")
        return None, None, None
